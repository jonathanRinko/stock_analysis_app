import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
import json
warnings.filterwarnings('ignore')

# For PDF export
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ ReportLab not installed. PDF export will not be available.")
    print("Install with: pip install reportlab")

# For Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ OpenPyXL not installed. Excel export will not be available.")
    print("Install with: pip install openpyxl")


class EnhancedStockValuationApp:
    """
    Comprehensive Stock Fundamental Valuation Application
    Enhanced with competitor analysis, historical trends, Monte Carlo, and export features
    """
    
    def __init__(self, ticker):
        """Initialize with stock ticker"""
        self.ticker = ticker.upper()
        self.stock = yf.Ticker(self.ticker)
        self.info = self.stock.info
        self.financials = None
        self.balance_sheet = None
        self.cashflow = None
        self.history = None
        self.competitors = []
        self.load_financial_data()
        
    def load_financial_data(self):
        """Load all financial statements and historical data"""
        try:
            self.financials = self.stock.financials
            self.balance_sheet = self.stock.balance_sheet
            self.cashflow = self.stock.cashflow
            self.history = self.stock.history(period="5y")
            print(f"✓ Successfully loaded data for {self.ticker}")
        except Exception as e:
            print(f"✗ Error loading data: {e}")
    
    def get_key_metrics(self):
        """Extract key financial metrics"""
        try:
            metrics = {
                'Company Name': self.info.get('longName', 'N/A'),
                'Sector': self.info.get('sector', 'N/A'),
                'Industry': self.info.get('industry', 'N/A'),
                'Market Cap': self.info.get('marketCap', 'N/A'),
                'Current Price': self.info.get('currentPrice', self.info.get('regularMarketPrice', 'N/A')),
                'PE Ratio (TTM)': self.info.get('trailingPE', 'N/A'),
                'Forward PE': self.info.get('forwardPE', 'N/A'),
                'PEG Ratio': self.info.get('pegRatio', 'N/A'),
                'Price to Book': self.info.get('priceToBook', 'N/A'),
                'Price to Sales': self.info.get('priceToSalesTrailing12Months', 'N/A'),
                'EV/EBITDA': self.info.get('enterpriseToEbitda', 'N/A'),
                'Profit Margin': self.info.get('profitMargins', 'N/A'),
                'Operating Margin': self.info.get('operatingMargins', 'N/A'),
                'ROE': self.info.get('returnOnEquity', 'N/A'),
                'ROA': self.info.get('returnOnAssets', 'N/A'),
                'Debt to Equity': self.info.get('debtToEquity', 'N/A'),
                'Current Ratio': self.info.get('currentRatio', 'N/A'),
                'Quick Ratio': self.info.get('quickRatio', 'N/A'),
                'Dividend Yield': self.info.get('dividendYield', 'N/A'),
                'Beta': self.info.get('beta', 'N/A'),
            }
            return metrics
        except Exception as e:
            print(f"Error getting metrics: {e}")
            return {}
    
    # ==================== COMPETITOR ANALYSIS ====================
    
    def get_competitors(self, custom_competitors=None):
        """
        Get competitor tickers for comparison
        """
        if custom_competitors:
            self.competitors = custom_competitors
        else:
            # Try to get from Yahoo Finance recommendations
            try:
                # Get similar stocks based on sector/industry
                sector = self.info.get('sector', '')
                industry = self.info.get('industry', '')
                
                # Common competitors by sector (expandable)
                competitor_map = {
                    'Technology': {
                        'AAPL': ['MSFT', 'GOOGL', 'META', 'AMZN'],
                        'MSFT': ['AAPL', 'GOOGL', 'ORCL', 'IBM'],
                        'GOOGL': ['META', 'MSFT', 'AMZN', 'AAPL'],
                        'NVDA': ['AMD', 'INTC', 'QCOM', 'AVGO'],
                        'AMD': ['NVDA', 'INTC', 'QCOM', 'MU'],
                    },
                    'Financial Services': {
                        'JPM': ['BAC', 'WFC', 'C', 'GS'],
                        'BAC': ['JPM', 'WFC', 'C', 'USB'],
                        'V': ['MA', 'AXP', 'PYPL', 'SQ'],
                    },
                    'Healthcare': {
                        'JNJ': ['PFE', 'MRK', 'ABBV', 'LLY'],
                        'UNH': ['ANTM', 'CI', 'HUM', 'CVS'],
                    },
                    'Consumer Cyclical': {
                        'AMZN': ['WMT', 'TGT', 'COST', 'HD'],
                        'TSLA': ['F', 'GM', 'TM', 'RIVN'],
                    }
                }
                
                # Try to find competitors
                for sec, comp_dict in competitor_map.items():
                    if sec in sector:
                        if self.ticker in comp_dict:
                            self.competitors = comp_dict[self.ticker]
                            break
                
                # If no competitors found, use generic approach
                if not self.competitors:
                    print(f"⚠️ No predefined competitors found for {self.ticker}")
                    self.competitors = []
                    
            except Exception as e:
                print(f"Error finding competitors: {e}")
                self.competitors = []
        
        return self.competitors
    
    def compare_with_competitors(self, competitors=None):
        """
        Comprehensive competitor comparison
        """
        if competitors:
            self.get_competitors(competitors)
        elif not self.competitors:
            self.get_competitors()
        
        if not self.competitors:
            return None
        
        comparison_data = []
        
        # Add main stock data
        main_data = {
            'Ticker': self.ticker,
            'Company': self.info.get('longName', self.ticker),
            'Market Cap': self.info.get('marketCap', 0),
            'P/E Ratio': self.info.get('trailingPE', None),
            'Forward P/E': self.info.get('forwardPE', None),
            'PEG Ratio': self.info.get('pegRatio', None),
            'Price/Book': self.info.get('priceToBook', None),
            'Price/Sales': self.info.get('priceToSalesTrailing12Months', None),
            'EV/EBITDA': self.info.get('enterpriseToEbitda', None),
            'Profit Margin': self.info.get('profitMargins', None),
            'ROE': self.info.get('returnOnEquity', None),
            'ROA': self.info.get('returnOnAssets', None),
            'Debt/Equity': self.info.get('debtToEquity', None),
            'Current Ratio': self.info.get('currentRatio', None),
            'Dividend Yield': self.info.get('dividendYield', None),
            'Revenue Growth': self.info.get('revenueGrowth', None),
            'Earnings Growth': self.info.get('earningsGrowth', None),
            'Beta': self.info.get('beta', None),
        }
        comparison_data.append(main_data)
        
        # Add competitor data
        for comp_ticker in self.competitors:
            try:
                comp_stock = yf.Ticker(comp_ticker)
                comp_info = comp_stock.info
                
                comp_data = {
                    'Ticker': comp_ticker,
                    'Company': comp_info.get('longName', comp_ticker),
                    'Market Cap': comp_info.get('marketCap', 0),
                    'P/E Ratio': comp_info.get('trailingPE', None),
                    'Forward P/E': comp_info.get('forwardPE', None),
                    'PEG Ratio': comp_info.get('pegRatio', None),
                    'Price/Book': comp_info.get('priceToBook', None),
                    'Price/Sales': comp_info.get('priceToSalesTrailing12Months', None),
                    'EV/EBITDA': comp_info.get('enterpriseToEbitda', None),
                    'Profit Margin': comp_info.get('profitMargins', None),
                    'ROE': comp_info.get('returnOnEquity', None),
                    'ROA': comp_info.get('returnOnAssets', None),
                    'Debt/Equity': comp_info.get('debtToEquity', None),
                    'Current Ratio': comp_info.get('currentRatio', None),
                    'Dividend Yield': comp_info.get('dividendYield', None),
                    'Revenue Growth': comp_info.get('revenueGrowth', None),
                    'Earnings Growth': comp_info.get('earningsGrowth', None),
                    'Beta': comp_info.get('beta', None),
                }
                comparison_data.append(comp_data)
                
            except Exception as e:
                print(f"Error loading competitor {comp_ticker}: {e}")
        
        # Create DataFrame
        df_comparison = pd.DataFrame(comparison_data)
        
        # Calculate industry averages
        numeric_cols = df_comparison.select_dtypes(include=[np.number]).columns
        industry_avg = df_comparison[numeric_cols].mean()
        industry_avg['Ticker'] = 'INDUSTRY AVG'
        industry_avg['Company'] = 'Industry Average'
        
        # Append industry average
        df_comparison = pd.concat([df_comparison, pd.DataFrame([industry_avg])], ignore_index=True)
        
        return df_comparison
    
    # ==================== HISTORICAL TREND ANALYSIS ====================
    
    def historical_trend_analysis(self, years=5):
        """
        Analyze historical trends in key metrics
        """
        try:
            # Get historical data
            end_date = datetime.now()
            start_date = end_date - timedelta(days=years*365)
            
            hist_data = self.stock.history(start=start_date, end=end_date)
            
            if hist_data.empty:
                return None
            
            # Calculate various metrics
            hist_data['Returns'] = hist_data['Close'].pct_change()
            hist_data['Cumulative_Returns'] = (1 + hist_data['Returns']).cumprod()
            hist_data['Volatility'] = hist_data['Returns'].rolling(window=30).std() * np.sqrt(252)
            
            # Moving averages
            hist_data['SMA_50'] = hist_data['Close'].rolling(window=50).mean()
            hist_data['SMA_200'] = hist_data['Close'].rolling(window=200).mean()
            
            # Price momentum
            hist_data['Momentum'] = hist_data['Close'] - hist_data['Close'].shift(10)
            
            # Calculate annual returns
            annual_returns = hist_data['Returns'].resample('Y').apply(lambda x: (1 + x).prod() - 1)
            
            # Historical metrics summary
            summary = {
                'Total Return': (hist_data['Close'].iloc[-1] / hist_data['Close'].iloc[0] - 1) * 100,
                'Annualized Return': ((hist_data['Close'].iloc[-1] / hist_data['Close'].iloc[0]) ** (1/years) - 1) * 100,
                'Volatility (Annual)': hist_data['Returns'].std() * np.sqrt(252) * 100,
                'Sharpe Ratio': (hist_data['Returns'].mean() / hist_data['Returns'].std() * np.sqrt(252)) if hist_data['Returns'].std() != 0 else 0,
                'Max Drawdown': ((hist_data['Close'] / hist_data['Close'].cummax()) - 1).min() * 100,
                'Best Year': annual_returns.max() * 100,
                'Worst Year': annual_returns.min() * 100,
                'Positive Years': (annual_returns > 0).sum(),
                'Negative Years': (annual_returns < 0).sum(),
                'Current Price': hist_data['Close'].iloc[-1],
                '52-Week High': hist_data['Close'].iloc[-252:].max() if len(hist_data) >= 252 else hist_data['Close'].max(),
                '52-Week Low': hist_data['Close'].iloc[-252:].min() if len(hist_data) >= 252 else hist_data['Close'].min(),
            }
            
            return {
                'historical_data': hist_data,
                'summary': summary,
                'annual_returns': annual_returns
            }
            
        except Exception as e:
            print(f"Error in historical analysis: {e}")
            return None
    
    def plot_historical_trends(self, save_path='charts/'):
        """
        Create visualization of historical trends
        """
        import os
        os.makedirs(save_path, exist_ok=True)
        
        trends = self.historical_trend_analysis()
        if not trends:
            return None
        
        hist_data = trends['historical_data']
        
        # Create figure with subplots
        fig, axes = plt.subplots(3, 2, figsize=(15, 12))
        fig.suptitle(f'{self.ticker} Historical Analysis', fontsize=16, fontweight='bold')
        
        # 1. Price and Moving Averages
        ax1 = axes[0, 0]
        ax1.plot(hist_data.index, hist_data['Close'], label='Close Price', linewidth=1.5)
        ax1.plot(hist_data.index, hist_data['SMA_50'], label='50-Day SMA', alpha=0.7)
        ax1.plot(hist_data.index, hist_data['SMA_200'], label='200-Day SMA', alpha=0.7)
        ax1.set_title('Price & Moving Averages')
        ax1.set_ylabel('Price ($)')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        
        # 2. Volume
        ax2 = axes[0, 1]
        ax2.bar(hist_data.index, hist_data['Volume'], alpha=0.6, color='blue')
        ax2.set_title('Trading Volume')
        ax2.set_ylabel('Volume')
        ax2.grid(True, alpha=0.3)
        
        # 3. Cumulative Returns
        ax3 = axes[1, 0]
        ax3.plot(hist_data.index, (hist_data['Cumulative_Returns'] - 1) * 100, color='green', linewidth=2)
        ax3.set_title('Cumulative Returns')
        ax3.set_ylabel('Return (%)')
        ax3.axhline(y=0, color='r', linestyle='--', alpha=0.5)
        ax3.grid(True, alpha=0.3)
        
        # 4. Volatility
        ax4 = axes[1, 1]
        ax4.plot(hist_data.index, hist_data['Volatility'] * 100, color='red', linewidth=1.5)
        ax4.set_title('Rolling 30-Day Volatility')
        ax4.set_ylabel('Volatility (%)')
        ax4.grid(True, alpha=0.3)
        
        # 5. Returns Distribution
        ax5 = axes[2, 0]
        hist_data['Returns'].dropna().hist(bins=50, ax=ax5, alpha=0.7, color='purple')
        ax5.set_title('Returns Distribution')
        ax5.set_xlabel('Daily Returns')
        ax5.set_ylabel('Frequency')
        ax5.axvline(x=0, color='r', linestyle='--', alpha=0.5)
        ax5.grid(True, alpha=0.3)
        
        # 6. Annual Returns
        ax6 = axes[2, 1]
        annual_returns = trends['annual_returns'] * 100
        colors_bar = ['green' if x > 0 else 'red' for x in annual_returns]
        ax6.bar(range(len(annual_returns)), annual_returns, color=colors_bar, alpha=0.7)
        ax6.set_title('Annual Returns')
        ax6.set_ylabel('Return (%)')
        ax6.set_xlabel('Year')
        ax6.set_xticks(range(len(annual_returns)))
        ax6.set_xticklabels([d.year for d in annual_returns.index], rotation=45)
        ax6.axhline(y=0, color='black', linestyle='--', alpha=0.5)
        ax6.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # Save the figure
        chart_path = os.path.join(save_path, f'{self.ticker}_historical_analysis.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        print(f"✓ Chart saved to {chart_path}")
        
        return chart_path
    
    # ==================== TECHNICAL INDICATORS ====================
    
    def calculate_technical_indicators(self):
        """
        Calculate comprehensive technical indicators
        """
        try:
            if self.history is None or self.history.empty:
                self.history = self.stock.history(period="1y")
            
            df = self.history.copy()
            
            # RSI (Relative Strength Index)
            def calculate_rsi(data, periods=14):
                delta = data.diff()
                gain = (delta.where(delta > 0, 0)).rolling(window=periods).mean()
                loss = (-delta.where(delta < 0, 0)).rolling(window=periods).mean()
                rs = gain / loss
                rsi = 100 - (100 / (1 + rs))
                return rsi
            
            df['RSI'] = calculate_rsi(df['Close'])
            
            # MACD (Moving Average Convergence Divergence)
            exp1 = df['Close'].ewm(span=12, adjust=False).mean()
            exp2 = df['Close'].ewm(span=26, adjust=False).mean()
            df['MACD'] = exp1 - exp2
            df['MACD_Signal'] = df['MACD'].ewm(span=9, adjust=False).mean()
            df['MACD_Histogram'] = df['MACD'] - df['MACD_Signal']
            
            # Bollinger Bands
            df['BB_Middle'] = df['Close'].rolling(window=20).mean()
            bb_std = df['Close'].rolling(window=20).std()
            df['BB_Upper'] = df['BB_Middle'] + (bb_std * 2)
            df['BB_Lower'] = df['BB_Middle'] - (bb_std * 2)
            df['BB_Width'] = ((df['BB_Upper'] - df['BB_Lower']) / df['BB_Middle']) * 100
            
            # Stochastic Oscillator
            low_14 = df['Low'].rolling(window=14).min()
            high_14 = df['High'].rolling(window=14).max()
            df['Stochastic_%K'] = ((df['Close'] - low_14) / (high_14 - low_14)) * 100
            df['Stochastic_%D'] = df['Stochastic_%K'].rolling(window=3).mean()
            
            # ATR (Average True Range)
            high_low = df['High'] - df['Low']
            high_close = np.abs(df['High'] - df['Close'].shift())
            low_close = np.abs(df['Low'] - df['Close'].shift())
            ranges = pd.concat([high_low, high_close, low_close], axis=1)
            true_range = np.max(ranges, axis=1)
            df['ATR'] = true_range.rolling(14).mean()
            
            # OBV (On-Balance Volume)
            df['OBV'] = (np.sign(df['Close'].diff()) * df['Volume']).fillna(0).cumsum()
            
            # Current values and signals
            latest = df.iloc[-1]
            
            signals = {
                'RSI': {
                    'value': latest['RSI'],
                    'signal': 'Oversold' if latest['RSI'] < 30 else 'Overbought' if latest['RSI'] > 70 else 'Neutral'
                },
                'MACD': {
                    'value': latest['MACD'],
                    'signal_line': latest['MACD_Signal'],
                    'histogram': latest['MACD_Histogram'],
                    'signal': 'Bullish' if latest['MACD'] > latest['MACD_Signal'] else 'Bearish'
                },
                'Bollinger_Bands': {
                    'upper': latest['BB_Upper'],
                    'middle': latest['BB_Middle'],
                    'lower': latest['BB_Lower'],
                    'width': latest['BB_Width'],
                    'signal': 'Overbought' if latest['Close'] > latest['BB_Upper'] else 'Oversold' if latest['Close'] < latest['BB_Lower'] else 'Normal'
                },
                'Stochastic': {
                    'K': latest['Stochastic_%K'],
                    'D': latest['Stochastic_%D'],
                    'signal': 'Oversold' if latest['Stochastic_%K'] < 20 else 'Overbought' if latest['Stochastic_%K'] > 80 else 'Neutral'
                },
                'ATR': {
                    'value': latest['ATR'],
                    'interpretation': 'High volatility' if latest['ATR'] > df['ATR'].mean() else 'Low volatility'
                },
                'Moving_Averages': {
                    'SMA_50': df['Close'].rolling(50).mean().iloc[-1],
                    'SMA_200': df['Close'].rolling(200).mean().iloc[-1],
                    'signal': 'Bullish' if df['Close'].rolling(50).mean().iloc[-1] > df['Close'].rolling(200).mean().iloc[-1] else 'Bearish'
                }
            }
            
            return {
                'data': df,
                'signals': signals
            }
            
        except Exception as e:
            print(f"Error calculating technical indicators: {e}")
            return None
    
    def plot_technical_indicators(self, save_path='charts/'):
        """
        Visualize technical indicators
        """
        import os
        os.makedirs(save_path, exist_ok=True)
        
        tech_data = self.calculate_technical_indicators()
        if not tech_data:
            return None
        
        df = tech_data['data']
        
        # Create figure with subplots
        fig, axes = plt.subplots(4, 1, figsize=(15, 12))
        fig.suptitle(f'{self.ticker} Technical Indicators', fontsize=16, fontweight='bold')
        
        # 1. Price with Bollinger Bands
        ax1 = axes[0]
        ax1.plot(df.index, df['Close'], label='Close Price', linewidth=2, color='black')
        ax1.plot(df.index, df['BB_Upper'], label='BB Upper', linestyle='--', alpha=0.7, color='red')
        ax1.plot(df.index, df['BB_Middle'], label='BB Middle', linestyle='--', alpha=0.7, color='blue')
        ax1.plot(df.index, df['BB_Lower'], label='BB Lower', linestyle='--', alpha=0.7, color='green')
        ax1.fill_between(df.index, df['BB_Lower'], df['BB_Upper'], alpha=0.1)
        ax1.set_title('Price & Bollinger Bands')
        ax1.set_ylabel('Price ($)')
        ax1.legend(loc='upper left')
        ax1.grid(True, alpha=0.3)
        
        # 2. RSI
        ax2 = axes[1]
        ax2.plot(df.index, df['RSI'], label='RSI', linewidth=2, color='purple')
        ax2.axhline(y=70, color='r', linestyle='--', alpha=0.5, label='Overbought (70)')
        ax2.axhline(y=30, color='g', linestyle='--', alpha=0.5, label='Oversold (30)')
        ax2.fill_between(df.index, 30, 70, alpha=0.1)
        ax2.set_title('Relative Strength Index (RSI)')
        ax2.set_ylabel('RSI')
        ax2.set_ylim(0, 100)
        ax2.legend(loc='upper left')
        ax2.grid(True, alpha=0.3)
        
        # 3. MACD
        ax3 = axes[2]
        ax3.plot(df.index, df['MACD'], label='MACD', linewidth=2, color='blue')
        ax3.plot(df.index, df['MACD_Signal'], label='Signal Line', linewidth=2, color='red')
        ax3.bar(df.index, df['MACD_Histogram'], label='Histogram', alpha=0.3, color='gray')
        ax3.axhline(y=0, color='black', linestyle='-', alpha=0.3)
        ax3.set_title('MACD (Moving Average Convergence Divergence)')
        ax3.set_ylabel('MACD')
        ax3.legend(loc='upper left')
        ax3.grid(True, alpha=0.3)
        
        # 4. Stochastic Oscillator
        ax4 = axes[3]
        ax4.plot(df.index, df['Stochastic_%K'], label='%K', linewidth=2, color='blue')
        ax4.plot(df.index, df['Stochastic_%D'], label='%D', linewidth=2, color='red')
        ax4.axhline(y=80, color='r', linestyle='--', alpha=0.5, label='Overbought (80)')
        ax4.axhline(y=20, color='g', linestyle='--', alpha=0.5, label='Oversold (20)')
        ax4.fill_between(df.index, 20, 80, alpha=0.1)
        ax4.set_title('Stochastic Oscillator')
        ax4.set_ylabel('Value')
        ax4.set_ylim(0, 100)
        ax4.legend(loc='upper left')
        ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # Save the figure
        chart_path = os.path.join(save_path, f'{self.ticker}_technical_indicators.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        print(f"✓ Technical indicators chart saved to {chart_path}")
        
        return chart_path
    
    # ==================== MONTE CARLO SIMULATION ====================
    
    def monte_carlo_simulation(self, days=252, simulations=1000):
        """
        Monte Carlo simulation for price prediction
        
        Parameters:
        - days: Number of days to simulate (default: 252 = 1 year)
        - simulations: Number of simulation paths
        """
        try:
            # Get historical data
            if self.history is None or self.history.empty:
                self.history = self.stock.history(period="2y")
            
            # Calculate daily returns
            returns = self.history['Close'].pct_change().dropna()
            
            # Calculate statistics
            mean_return = returns.mean()
            std_return = returns.std()
            
            # Current price
            last_price = self.history['Close'].iloc[-1]
            
            # Run simulations
            simulation_results = np.zeros((days, simulations))
            
            for i in range(simulations):
                # Generate random returns
                daily_returns = np.random.normal(mean_return, std_return, days)
                
                # Calculate price path
                price_path = [last_price]
                for ret in daily_returns:
                    price_path.append(price_path[-1] * (1 + ret))
                
                simulation_results[:, i] = price_path[1:]
            
            # Calculate statistics
            final_prices = simulation_results[-1, :]
            
            percentiles = {
                '5th': np.percentile(final_prices, 5),
                '25th': np.percentile(final_prices, 25),
                '50th': np.percentile(final_prices, 50),
                '75th': np.percentile(final_prices, 75),
                '95th': np.percentile(final_prices, 95),
            }
            
            results = {
                'current_price': last_price,
                'simulation_days': days,
                'num_simulations': simulations,
                'mean_final_price': final_prices.mean(),
                'median_final_price': np.median(final_prices),
                'std_final_price': final_prices.std(),
                'min_final_price': final_prices.min(),
                'max_final_price': final_prices.max(),
                'percentiles': percentiles,
                'expected_return': ((final_prices.mean() - last_price) / last_price) * 100,
                'probability_profit': (final_prices > last_price).sum() / simulations * 100,
                'value_at_risk_5': percentiles['5th'],
                'simulation_paths': simulation_results,
                'final_prices': final_prices
            }
            
            return results
            
        except Exception as e:
            print(f"Error in Monte Carlo simulation: {e}")
            return None
    
    def plot_monte_carlo(self, mc_results=None, save_path='charts/'):
        """
        Visualize Monte Carlo simulation results
        """
        import os
        os.makedirs(save_path, exist_ok=True)
        
        if mc_results is None:
            mc_results = self.monte_carlo_simulation()
        
        if not mc_results:
            return None
        
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle(f'{self.ticker} Monte Carlo Simulation ({mc_results["num_simulations"]} simulations, {mc_results["simulation_days"]} days)', 
                     fontsize=16, fontweight='bold')
        
        # 1. Simulation paths
        ax1 = axes[0, 0]
        simulation_paths = mc_results['simulation_paths']
        # Plot only subset of paths for clarity
        for i in range(0, mc_results['num_simulations'], max(1, mc_results['num_simulations']//100)):
            ax1.plot(simulation_paths[:, i], alpha=0.1, color='blue', linewidth=0.5)
        
        # Plot percentiles
        ax1.plot(np.percentile(simulation_paths, 50, axis=1), color='red', linewidth=2, label='Median')
        ax1.plot(np.percentile(simulation_paths, 5, axis=1), color='green', linewidth=1.5, linestyle='--', label='5th Percentile')
        ax1.plot(np.percentile(simulation_paths, 95, axis=1), color='orange', linewidth=1.5, linestyle='--', label='95th Percentile')
        ax1.axhline(y=mc_results['current_price'], color='black', linestyle='-', linewidth=2, label='Current Price')
        ax1.set_title('Price Simulation Paths')
        ax1.set_xlabel('Days')
        ax1.set_ylabel('Price ($)')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        
        # 2. Distribution of final prices
        ax2 = axes[0, 1]
        final_prices = mc_results['final_prices']
        ax2.hist(final_prices, bins=50, alpha=0.7, color='skyblue', edgecolor='black')
        ax2.axvline(mc_results['current_price'], color='black', linestyle='--', linewidth=2, label='Current Price')
        ax2.axvline(mc_results['mean_final_price'], color='red', linestyle='--', linewidth=2, label='Mean Prediction')
        ax2.axvline(mc_results['percentiles']['5th'], color='green', linestyle='--', linewidth=1.5, label='5th Percentile')
        ax2.axvline(mc_results['percentiles']['95th'], color='orange', linestyle='--', linewidth=1.5, label='95th Percentile')
        ax2.set_title('Distribution of Final Prices')
        ax2.set_xlabel('Price ($)')
        ax2.set_ylabel('Frequency')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        # 3. Percentile bands
        ax3 = axes[1, 0]
        percentile_5 = np.percentile(simulation_paths, 5, axis=1)
        percentile_25 = np.percentile(simulation_paths, 25, axis=1)
        percentile_50 = np.percentile(simulation_paths, 50, axis=1)
        percentile_75 = np.percentile(simulation_paths, 75, axis=1)
        percentile_95 = np.percentile(simulation_paths, 95, axis=1)
        
        days_range = range(len(percentile_50))
        ax3.fill_between(days_range, percentile_5, percentile_95, alpha=0.2, color='blue', label='5-95 Percentile')
        ax3.fill_between(days_range, percentile_25, percentile_75, alpha=0.3, color='green', label='25-75 Percentile')
        ax3.plot(percentile_50, color='red', linewidth=2, label='Median')
        ax3.axhline(y=mc_results['current_price'], color='black', linestyle='--', linewidth=1.5, label='Current Price')
        ax3.set_title('Confidence Bands')
        ax3.set_xlabel('Days')
        ax3.set_ylabel('Price ($)')
        ax3.legend()
        ax3.grid(True, alpha=0.3)
        
        # 4. Statistics summary (text)
        ax4 = axes[1, 1]
        ax4.axis('off')
        
        stats_text = f"""
        MONTE CARLO STATISTICS
        {'='*40}
        
        Current Price:           ${mc_results['current_price']:.2f}
        
        Expected Final Price:    ${mc_results['mean_final_price']:.2f}
        Median Final Price:      ${mc_results['median_final_price']:.2f}
        
        Expected Return:         {mc_results['expected_return']:.2f}%
        Probability of Profit:   {mc_results['probability_profit']:.1f}%
        
        PRICE RANGE
        {'='*40}
        Minimum:                 ${mc_results['min_final_price']:.2f}
        5th Percentile:          ${mc_results['percentiles']['5th']:.2f}
        25th Percentile:         ${mc_results['percentiles']['25th']:.2f}
        50th Percentile:         ${mc_results['percentiles']['50th']:.2f}
        75th Percentile:         ${mc_results['percentiles']['75th']:.2f}
        95th Percentile:         ${mc_results['percentiles']['95th']:.2f}
        Maximum:                 ${mc_results['max_final_price']:.2f}
        
        RISK METRICS
        {'='*40}
        Value at Risk (5%):      ${mc_results['value_at_risk_5']:.2f}
        Standard Deviation:      ${mc_results['std_final_price']:.2f}
        """
        
        ax4.text(0.1, 0.95, stats_text, transform=ax4.transAxes, 
                fontfamily='monospace', fontsize=10, verticalalignment='top')
        
        plt.tight_layout()
        
        # Save the figure
        chart_path = os.path.join(save_path, f'{self.ticker}_monte_carlo.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        print(f"✓ Monte Carlo chart saved to {chart_path}")
        
        return chart_path
    
    # ==================== EXISTING VALUATION METHODS ====================
    
    def dcf_valuation(self, growth_rate=0.05, terminal_growth=0.02, wacc=0.10, years=5):
        """Discounted Cash Flow (DCF) Valuation"""
        try:
            if 'Free Cash Flow' in self.cashflow.index:
                fcf = self.cashflow.loc['Free Cash Flow'].iloc[0]
            else:
                operating_cf = self.cashflow.loc['Total Cash From Operating Activities'].iloc[0]
                capex = self.cashflow.loc['Capital Expenditures'].iloc[0] if 'Capital Expenditures' in self.cashflow.index else 0
                fcf = operating_cf + capex
            
            projected_fcf = []
            for year in range(1, years + 1):
                future_fcf = fcf * ((1 + growth_rate) ** year)
                pv_fcf = future_fcf / ((1 + wacc) ** year)
                projected_fcf.append(pv_fcf)
            
            terminal_fcf = fcf * ((1 + growth_rate) ** years) * (1 + terminal_growth)
            terminal_value = terminal_fcf / (wacc - terminal_growth)
            pv_terminal_value = terminal_value / ((1 + wacc) ** years)
            
            enterprise_value = sum(projected_fcf) + pv_terminal_value
            
            total_debt = self.info.get('totalDebt', 0)
            cash = self.info.get('totalCash', 0)
            equity_value = enterprise_value - total_debt + cash
            
            shares_outstanding = self.info.get('sharesOutstanding', 1)
            intrinsic_value_per_share = equity_value / shares_outstanding
            
            current_price = self.info.get('currentPrice', self.info.get('regularMarketPrice', 0))
            upside = ((intrinsic_value_per_share - current_price) / current_price * 100) if current_price else 0
            
            return {
                'Free Cash Flow (Current)': fcf,
                'PV of Projected FCF': sum(projected_fcf),
                'Terminal Value': terminal_value,
                'PV of Terminal Value': pv_terminal_value,
                'Enterprise Value': enterprise_value,
                'Equity Value': equity_value,
                'Intrinsic Value per Share': intrinsic_value_per_share,
                'Current Price': current_price,
                'Upside/Downside %': upside
            }
        except Exception as e:
            print(f"Error in DCF calculation: {e}")
            return None
    
    def dividend_discount_model(self, growth_rate=0.05, required_return=0.10):
        """Gordon Growth Model (Dividend Discount Model)"""
        try:
            dividend_rate = self.info.get('dividendRate', 0)
            
            if dividend_rate == 0:
                return {'Error': 'Company does not pay dividends'}
            
            next_dividend = dividend_rate * (1 + growth_rate)
            
            if required_return <= growth_rate:
                return {'Error': 'Required return must be greater than growth rate'}
            
            intrinsic_value = next_dividend / (required_return - growth_rate)
            current_price = self.info.get('currentPrice', self.info.get('regularMarketPrice', 0))
            upside = ((intrinsic_value - current_price) / current_price * 100) if current_price else 0
            
            return {
                'Current Dividend': dividend_rate,
                'Expected Next Dividend': next_dividend,
                'Growth Rate': growth_rate,
                'Required Return': required_return,
                'Intrinsic Value': intrinsic_value,
                'Current Price': current_price,
                'Upside/Downside %': upside
            }
        except Exception as e:
            print(f"Error in DDM calculation: {e}")
            return None
    
    def pe_valuation(self, industry_pe=None):
        """P/E Ratio Valuation"""
        try:
            eps = self.info.get('trailingEps', 0)
            current_pe = self.info.get('trailingPE', 0)
            current_price = self.info.get('currentPrice', self.info.get('regularMarketPrice', 0))
            
            if not industry_pe:
                industry_pe = self.info.get('forwardPE', current_pe)
            
            intrinsic_value = eps * industry_pe if eps and industry_pe else 0
            upside = ((intrinsic_value - current_price) / current_price * 100) if current_price and intrinsic_value else 0
            
            return {
                'EPS (TTM)': eps,
                'Current P/E': current_pe,
                'Industry/Target P/E': industry_pe,
                'Intrinsic Value': intrinsic_value,
                'Current Price': current_price,
                'Upside/Downside %': upside
            }
        except Exception as e:
            print(f"Error in P/E valuation: {e}")
            return None
    
    def book_value_valuation(self):
        """Price to Book Valuation"""
        try:
            book_value = self.info.get('bookValue', 0)
            pb_ratio = self.info.get('priceToBook', 0)
            current_price = self.info.get('currentPrice', self.info.get('regularMarketPrice', 0))
            
            discount_to_book = ((current_price - book_value) / book_value * 100) if book_value else 0
            
            return {
                'Book Value per Share': book_value,
                'Current Price': current_price,
                'Price to Book Ratio': pb_ratio,
                'Premium/Discount to Book %': discount_to_book,
                'Interpretation': 'Undervalued' if pb_ratio < 1 else 'Overvalued' if pb_ratio > 3 else 'Fairly Valued'
            }
        except Exception as e:
            print(f"Error in Book Value calculation: {e}")
            return None
    
    def graham_number(self):
        """Benjamin Graham's Number"""
        try:
            eps = self.info.get('trailingEps', 0)
            book_value = self.info.get('bookValue', 0)
            
            if eps > 0 and book_value > 0:
                graham_number = (22.5 * eps * book_value) ** 0.5
                current_price = self.info.get('currentPrice', self.info.get('regularMarketPrice', 0))
                margin_of_safety = ((graham_number - current_price) / graham_number * 100) if graham_number else 0
                
                return {
                    'EPS': eps,
                    'Book Value': book_value,
                    'Graham Number': graham_number,
                    'Current Price': current_price,
                    'Margin of Safety %': margin_of_safety,
                    'Recommendation': 'Buy' if margin_of_safety > 20 else 'Hold' if margin_of_safety > 0 else 'Overvalued'
                }
            else:
                return {'Error': 'Negative or zero EPS/Book Value'}
        except Exception as e:
            print(f"Error in Graham Number calculation: {e}")
            return None
    
    def calculate_wacc(self):
        """Calculate Weighted Average Cost of Capital (WACC)"""
        try:
            market_cap = self.info.get('marketCap', 0)
            total_debt = self.info.get('totalDebt', 0)
            
            beta = self.info.get('beta', 1)
            risk_free_rate = 0.04
            market_return = 0.10
            cost_of_equity = risk_free_rate + beta * (market_return - risk_free_rate)
            
            interest_expense = abs(self.cashflow.loc['Interest Expense'].iloc[0]) if 'Interest Expense' in self.cashflow.index else 0
            cost_of_debt = (interest_expense / total_debt) if total_debt > 0 else 0
            
            tax_rate = self.info.get('taxRate', 0.21)
            
            total_value = market_cap + total_debt
            if total_value > 0:
                wacc = (market_cap / total_value) * cost_of_equity + \
                       (total_debt / total_value) * cost_of_debt * (1 - tax_rate)
            else:
                wacc = cost_of_equity
            
            return {
                'Cost of Equity': cost_of_equity,
                'Cost of Debt': cost_of_debt,
                'Tax Rate': tax_rate,
                'WACC': wacc,
                'Beta': beta
            }
        except Exception as e:
            print(f"Error calculating WACC: {e}")
            return None
    
    def financial_health_score(self):
        """Calculate overall financial health score (0-100)"""
        score = 0
        max_score = 100
        
        try:
            roe = self.info.get('returnOnEquity', 0)
            if roe and roe > 0.15:
                score += 15
            elif roe and roe > 0.10:
                score += 10
            elif roe and roe > 0.05:
                score += 5
            
            profit_margin = self.info.get('profitMargins', 0)
            if profit_margin and profit_margin > 0.20:
                score += 15
            elif profit_margin and profit_margin > 0.10:
                score += 10
            elif profit_margin and profit_margin > 0.05:
                score += 5
            
            current_ratio = self.info.get('currentRatio', 0)
            if current_ratio and current_ratio > 2:
                score += 10
            elif current_ratio and current_ratio > 1.5:
                score += 7
            elif current_ratio and current_ratio > 1:
                score += 4
            
            quick_ratio = self.info.get('quickRatio', 0)
            if quick_ratio and quick_ratio > 1.5:
                score += 10
            elif quick_ratio and quick_ratio > 1:
                score += 7
            elif quick_ratio and quick_ratio > 0.5:
                score += 4
            
            debt_to_equity = self.info.get('debtToEquity', 0)
            if debt_to_equity is not None:
                if debt_to_equity < 0.5:
                    score += 20
                elif debt_to_equity < 1:
                    score += 15
                elif debt_to_equity < 2:
                    score += 10
                elif debt_to_equity < 3:
                    score += 5
            
            pe_ratio = self.info.get('trailingPE', 0)
            if pe_ratio and 10 < pe_ratio < 20:
                score += 8
            elif pe_ratio and 5 < pe_ratio < 25:
                score += 5
            
            peg_ratio = self.info.get('pegRatio', 0)
            if peg_ratio and peg_ratio < 1:
                score += 7
            elif peg_ratio and peg_ratio < 2:
                score += 4
            
            revenue_growth = self.info.get('revenueGrowth', 0)
            if revenue_growth and revenue_growth > 0.20:
                score += 8
            elif revenue_growth and revenue_growth > 0.10:
                score += 5
            elif revenue_growth and revenue_growth > 0:
                score += 3
            
            earnings_growth = self.info.get('earningsGrowth', 0)
            if earnings_growth and earnings_growth > 0.20:
                score += 7
            elif earnings_growth and earnings_growth > 0.10:
                score += 4
            elif earnings_growth and earnings_growth > 0:
                score += 2
            
            if score >= 80:
                rating = "Excellent"
            elif score >= 60:
                rating = "Good"
            elif score >= 40:
                rating = "Fair"
            else:
                rating = "Poor"
            
            return {
                'Financial Health Score': f"{score}/{max_score}",
                'Rating': rating,
                'Percentage': f"{(score/max_score)*100:.1f}%"
            }
        except Exception as e:
            print(f"Error calculating health score: {e}")
            return None
    
    # ==================== EXPORT FUNCTIONS ====================
    
    def export_to_excel(self, filename=None):
        """
        Export comprehensive analysis to Excel
        """
        if not OPENPYXL_AVAILABLE:
            print("❌ OpenPyXL not installed. Cannot export to Excel.")
            return None
        
        if filename is None:
            filename = f"{self.ticker}_valuation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            subheader_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
            subheader_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. Summary Sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary['A1'] = f"{self.ticker} - Fundamental Valuation Report"
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            row = 4
            metrics = self.get_key_metrics()
            ws_summary.cell(row, 1, "Key Metrics").font = header_font
            ws_summary.cell(row, 1).fill = header_fill
            ws_summary.cell(row, 2).fill = header_fill
            row += 1
            
            for key, value in metrics.items():
                ws_summary.cell(row, 1, key)
                ws_summary.cell(row, 2, value)
                row += 1
            
            # Auto-adjust column widths
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20
            
            # 2. Valuation Models Sheet
            ws_valuation = wb.create_sheet("Valuation Models")
            row = 1
            
            # DCF
            ws_valuation.cell(row, 1, "DCF Valuation").font = header_font
            ws_valuation.cell(row, 1).fill = header_fill
            row += 1
            dcf = self.dcf_valuation()
            if dcf:
                for key, value in dcf.items():
                    ws_valuation.cell(row, 1, key)
                    ws_valuation.cell(row, 2, value)
                    row += 1
            row += 2
            
            # DDM
            ws_valuation.cell(row, 1, "Dividend Discount Model").font = header_font
            ws_valuation.cell(row, 1).fill = header_fill
            row += 1
            ddm = self.dividend_discount_model()
            if ddm and 'Error' not in ddm:
                for key, value in ddm.items():
                    ws_valuation.cell(row, 1, key)
                    ws_valuation.cell(row, 2, value)
                    row += 1
            row += 2
            
            # Graham Number
            ws_valuation.cell(row, 1, "Graham Number").font = header_font
            ws_valuation.cell(row, 1).fill = header_fill
            row += 1
            graham = self.graham_number()
            if graham and 'Error' not in graham:
                for key, value in graham.items():
                    ws_valuation.cell(row, 1, key)
                    ws_valuation.cell(row, 2, value)
                    row += 1
            
            ws_valuation.column_dimensions['A'].width = 30
            ws_valuation.column_dimensions['B'].width = 20
            
            # 3. Competitor Analysis Sheet
            ws_competitors = wb.create_sheet("Competitor Analysis")
            comp_df = self.compare_with_competitors()
            if comp_df is not None:
                for r_idx, row_data in enumerate(dataframe_to_rows(comp_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row_data, 1):
                        cell = ws_competitors.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                        cell.border = border
                
                # Auto-adjust column widths
                for column in ws_competitors.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_competitors.column_dimensions[column_letter].width = adjusted_width
            
            # 4. Historical Analysis Sheet
            ws_historical = wb.create_sheet("Historical Analysis")
            hist_analysis = self.historical_trend_analysis()
            if hist_analysis:
                ws_historical.cell(1, 1, "Historical Performance Summary").font = header_font
                ws_historical.cell(1, 1).fill = header_fill
                row = 2
                for key, value in hist_analysis['summary'].items():
                    ws_historical.cell(row, 1, key)
                    ws_historical.cell(row, 2, value)
                    row += 1
            
            ws_historical.column_dimensions['A'].width = 30
            ws_historical.column_dimensions['B'].width = 20
            
            # 5. Technical Indicators Sheet
            ws_technical = wb.create_sheet("Technical Indicators")
            tech_data = self.calculate_technical_indicators()
            if tech_data:
                ws_technical.cell(1, 1, "Technical Indicator Signals").font = header_font
                ws_technical.cell(1, 1).fill = header_fill
                row = 2
                for indicator, data in tech_data['signals'].items():
                    ws_technical.cell(row, 1, indicator).font = subheader_font
                    ws_technical.cell(row, 1).fill = subheader_fill
                    row += 1
                    for key, value in data.items():
                        ws_technical.cell(row, 1, f"  {key}")
                        ws_technical.cell(row, 2, value)
                        row += 1
                    row += 1
            
            ws_technical.column_dimensions['A'].width = 30
            ws_technical.column_dimensions['B'].width = 20
            
            # 6. Monte Carlo Sheet
            ws_mc = wb.create_sheet("Monte Carlo Simulation")
            mc_results = self.monte_carlo_simulation()
            if mc_results:
                ws_mc.cell(1, 1, "Monte Carlo Simulation Results").font = header_font
                ws_mc.cell(1, 1).fill = header_fill
                row = 2
                
                # Add key results
                for key, value in mc_results.items():
                    if key not in ['simulation_paths', 'final_prices']:
                        if key == 'percentiles':
                            ws_mc.cell(row, 1, "Percentiles").font = subheader_font
                            row += 1
                            for pct_key, pct_value in value.items():
                                ws_mc.cell(row, 1, f"  {pct_key}")
                                ws_mc.cell(row, 2, pct_value)
                                row += 1
                        else:
                            ws_mc.cell(row, 1, key)
                            ws_mc.cell(row, 2, value)
                            row += 1
            
            ws_mc.column_dimensions['A'].width = 30
            ws_mc.column_dimensions['B'].width = 20
            
            # Save workbook
            wb.save(filename)
            print(f"✓ Excel report exported to {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            return None
    
    def export_to_pdf(self, filename=None, include_charts=True):
        """
        Export comprehensive analysis to PDF
        """
        if not REPORTLAB_AVAILABLE:
            print("❌ ReportLab not installed. Cannot export to PDF.")
            return None
        
        if filename is None:
            filename = f"{self.ticker}_valuation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=letter,
                                  topMargin=0.75*inch, bottomMargin=0.75*inch)
            
            # Container for elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=30,
                alignment=1  # Center
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            # Title
            elements.append(Paragraph(f"{self.ticker} - Fundamental Valuation Report", title_style))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Key Metrics
            elements.append(Paragraph("Key Financial Metrics", heading_style))
            metrics = self.get_key_metrics()
            metrics_data = [['Metric', 'Value']]
            for key, value in list(metrics.items())[:15]:  # Limit to first 15
                metrics_data.append([key, str(value)])
            
            metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(metrics_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # DCF Valuation
            elements.append(Paragraph("DCF Valuation", heading_style))
            dcf = self.dcf_valuation()
            if dcf:
                dcf_data = [['Metric', 'Value']]
                for key, value in dcf.items():
                    if isinstance(value, (int, float)):
                        dcf_data.append([key, f"${value:,.2f}" if 'Price' in key or 'Value' in key else f"{value:.2f}"])
                    else:
                        dcf_data.append([key, str(value)])
                
                dcf_table = Table(dcf_data, colWidths=[3*inch, 2*inch])
                dcf_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(dcf_table)
            elements.append(PageBreak())
            
            # Competitor Analysis
            elements.append(Paragraph("Competitor Analysis", heading_style))
            comp_df = self.compare_with_competitors()
            if comp_df is not None:
                # Select key columns for PDF
                key_cols = ['Ticker', 'Company', 'Market Cap', 'P/E Ratio', 'ROE', 'Debt/Equity']
                comp_subset = comp_df[key_cols] if all(col in comp_df.columns for col in key_cols) else comp_df.iloc[:, :6]
                
                comp_data = [comp_subset.columns.tolist()]
                for _, row in comp_subset.iterrows():
                    comp_data.append([str(val)[:20] for val in row.tolist()])  # Truncate long values
                
                comp_table = Table(comp_data, colWidths=[1.2*inch]*len(comp_subset.columns))
                comp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(comp_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Historical Performance
            elements.append(Paragraph("Historical Performance Summary", heading_style))
            hist_analysis = self.historical_trend_analysis()
            if hist_analysis:
                hist_data = [['Metric', 'Value']]
                for key, value in hist_analysis['summary'].items():
                    if isinstance(value, (int, float)):
                        hist_data.append([key, f"{value:.2f}"])
                    else:
                        hist_data.append([key, str(value)])
                
                hist_table = Table(hist_data, colWidths=[3*inch, 2*inch])
                hist_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(hist_table)
            
            # Add charts if requested
            if include_charts:
                elements.append(PageBreak())
                elements.append(Paragraph("Charts & Visualizations", heading_style))
                
                # Generate and add charts
                try:
                    chart1 = self.plot_historical_trends()
                    if chart1:
                        elements.append(Image(chart1, width=6*inch, height=4.5*inch))
                        elements.append(Spacer(1, 0.2*inch))
                except Exception as e:
                    print(f"Warning: Could not add historical chart: {e}")
                
                try:
                    chart2 = self.plot_technical_indicators()
                    if chart2:
                        elements.append(PageBreak())
                        elements.append(Image(chart2, width=6*inch, height=4.5*inch))
                        elements.append(Spacer(1, 0.2*inch))
                except Exception as e:
                    print(f"Warning: Could not add technical chart: {e}")
                
                try:
                    chart3 = self.plot_monte_carlo()
                    if chart3:
                        elements.append(PageBreak())
                        elements.append(Image(chart3, width=6*inch, height=4*inch))
                except Exception as e:
                    print(f"Warning: Could not add Monte Carlo chart: {e}")
            
            # Disclaimer
            elements.append(PageBreak())
            disclaimer_text = """
            <b>DISCLAIMER</b><br/>
            This analysis is for educational and informational purposes only. 
            It does not constitute financial advice, investment recommendation, 
            or an offer to buy or sell securities. Always conduct your own research 
            and consult with a qualified financial advisor before making investment decisions.
            Past performance does not guarantee future results.
            """
            elements.append(Paragraph(disclaimer_text, styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            print(f"✓ PDF report exported to {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error exporting to PDF: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_full_report(self, export_excel=False, export_pdf=False, include_charts=True):
        """Generate comprehensive valuation report with optional exports"""
        print(f"\n{'='*80}")
        print(f"ENHANCED FUNDAMENTAL VALUATION REPORT: {self.ticker}")
        print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*80}\n")
        
        # Key Metrics
        print("📊 KEY FINANCIAL METRICS")
        print("-" * 80)
        metrics = self.get_key_metrics()
        for key, value in metrics.items():
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if value > 1000000:
                    print(f"{key:.<40} ${value:,.0f}")
                elif key in ['PE Ratio (TTM)', 'Forward PE', 'PEG Ratio', 'Price to Book', 
                            'Price to Sales', 'EV/EBITDA', 'Debt to Equity', 'Beta']:
                    print(f"{key:.<40} {value:.2f}")
                elif key in ['Profit Margin', 'Operating Margin', 'ROE', 'ROA', 'Dividend Yield']:
                    print(f"{key:.<40} {value*100:.2f}%" if value != 'N/A' else f"{key:.<40} N/A")
                else:
                    print(f"{key:.<40} {value:.2f}")
            else:
                print(f"{key:.<40} {value}")
        
        # DCF Valuation
        print(f"\n\n💰 DISCOUNTED CASH FLOW (DCF) VALUATION")
        print("-" * 80)
        dcf = self.dcf_valuation()
        if dcf:
            for key, value in dcf.items():
                if isinstance(value, (int, float)):
                    if 'Price' in key or 'Value' in key or 'FCF' in key or 'PV' in key:
                        print(f"{key:.<40} ${value:,.2f}")
                    elif '%' in key:
                        print(f"{key:.<40} {value:.2f}%")
                    else:
                        print(f"{key:.<40} {value:,.2f}")
        
        # Financial Health Score
        print(f"\n\n🏥 FINANCIAL HEALTH SCORE")
        print("-" * 80)
        health = self.financial_health_score()
        if health:
            for key, value in health.items():
                print(f"{key:.<40} {value}")
        
        # Competitor Analysis
        print(f"\n\n🔍 COMPETITOR ANALYSIS")
        print("-" * 80)
        comp_df = self.compare_with_competitors()
        if comp_df is not None:
            print("\nTop competitors comparison:")
            # Display subset of columns for readability
            display_cols = ['Ticker', 'Company', 'Market Cap', 'P/E Ratio', 'ROE', 'Profit Margin']
            available_cols = [col for col in display_cols if col in comp_df.columns]
            print(comp_df[available_cols].to_string(index=False))
        else:
            print("No competitor data available")
        
        # Historical Analysis
        print(f"\n\n📈 HISTORICAL PERFORMANCE")
        print("-" * 80)
        hist_analysis = self.historical_trend_analysis()
        if hist_analysis:
            for key, value in hist_analysis['summary'].items():
                if isinstance(value, (int, float)):
                    print(f"{key:.<40} {value:.2f}")
                else:
                    print(f"{key:.<40} {value}")
        
        # Technical Indicators
        print(f"\n\n📉 TECHNICAL INDICATORS")
        print("-" * 80)
        tech_data = self.calculate_technical_indicators()
        if tech_data:
            signals = tech_data['signals']
            for indicator, data in signals.items():
                print(f"\n{indicator}:")
                for key, value in data.items():
                    if isinstance(value, (int, float)):
                        print(f"  {key:.<35} {value:.2f}")
                    else:
                        print(f"  {key:.<35} {value}")
        
        # Monte Carlo Simulation
        print(f"\n\n🎲 MONTE CARLO SIMULATION")
        print("-" * 80)
        mc_results = self.monte_carlo_simulation()
        if mc_results:
            print(f"Current Price:                         ${mc_results['current_price']:.2f}")
            print(f"Expected Price ({mc_results['simulation_days']} days):          ${mc_results['mean_final_price']:.2f}")
            print(f"Expected Return:                       {mc_results['expected_return']:.2f}%")
            print(f"Probability of Profit:                 {mc_results['probability_profit']:.1f}%")
            print(f"\nPrice Range (95% Confidence):")
            print(f"  Lower Bound (5th percentile):        ${mc_results['percentiles']['5th']:.2f}")
            print(f"  Upper Bound (95th percentile):       ${mc_results['percentiles']['95th']:.2f}")
        
        # Generate charts if requested
        if include_charts:
            print(f"\n\n📊 GENERATING CHARTS...")
            print("-" * 80)
            self.plot_historical_trends()
            self.plot_technical_indicators()
            self.plot_monte_carlo(mc_results)
        
        # Export options
        if export_excel:
            print(f"\n\n📑 EXPORTING TO EXCEL...")
            print("-" * 80)
            self.export_to_excel()
        
        if export_pdf:
            print(f"\n\n📄 EXPORTING TO PDF...")
            print("-" * 80)
            self.export_to_pdf(include_charts=include_charts)
        
        print(f"\n{'='*80}")
        print("DISCLAIMER: This analysis is for educational purposes only.")
        print("Always conduct your own research before making investment decisions.")
        print(f"{'='*80}\n")

# Interactive Interface
def main():
    """Main function to run the enhanced valuation app"""
    print("\n" + "="*80)
    print(" "*15 + "ENHANCED STOCK FUNDAMENTAL VALUATION APP")
    print(" "*20 + "Advanced Quantitative Analysis Tool")
    print("="*80 + "\n")
    print("Features:")
    print("  ✓ DCF & Multiple Valuation Models")
    print("  ✓ Competitor Comparison")
    print("  ✓ Historical Trend Analysis")
    print("  ✓ Technical Indicators")
    print("  ✓ Monte Carlo Simulation")
    print("  ✓ Excel & PDF Export")
    print("="*80 + "\n")
    
    while True:
        ticker = input("Enter stock ticker (or 'quit' to exit): ").strip().upper()
        
        if ticker.lower() == 'quit':
            print("\nThank you for using the Enhanced Stock Valuation App!")
            break
        
        if not ticker:
            print("Please enter a valid ticker symbol.")
            continue
        
        try:
            # Create valuation instance
            print(f"\n🔄 Loading data for {ticker}...")
            valuation = EnhancedStockValuationApp(ticker)
            
            # Ask for competitor tickers
            comp_input = input("\nEnter competitor tickers (comma-separated, or press Enter to skip): ").strip()
            if comp_input:
                competitors = [c.strip().upper() for c in comp_input.split(',')]
                valuation.get_competitors(competitors)
            else:
                valuation.get_competitors()
            
            # Ask for export options
            print("\n" + "-"*80)
            print("EXPORT OPTIONS")
            print("-"*80)
            export_excel = input("Export to Excel? (yes/no, default: no): ").strip().lower() in ['yes', 'y']
            export_pdf = input("Export to PDF? (yes/no, default: no): ").strip().lower() in ['yes', 'y']
            include_charts = input("Include charts? (yes/no, default: yes): ").strip().lower() not in ['no', 'n']
            
            # Generate full report
            print("\n" + "="*80)
            print("GENERATING COMPREHENSIVE ANALYSIS...")
            print("="*80)
            
            valuation.generate_full_report(
                export_excel=export_excel,
                export_pdf=export_pdf,
                include_charts=include_charts
            )
            
            # Ask if user wants to analyze another stock
            print("\n" + "-"*80)
            another = input("\nAnalyze another stock? (yes/no): ").strip().lower()
            if another not in ['yes', 'y']:
                print("\nThank you for using the Enhanced Stock Valuation App!")
                break
                
        except KeyboardInterrupt:
            print("\n\n⚠️  Analysis interrupted by user.")
            print("\nThank you for using the Enhanced Stock Valuation App!")
            break
            
        except Exception as e:
            print(f"\n❌ Error analyzing {ticker}: {e}")
            print("Please check the ticker symbol and try again.\n")
            import traceback
            traceback.print_exc()
            
            # Ask if user wants to try another ticker
            retry = input("\nTry another ticker? (yes/no): ").strip().lower()
            if retry not in ['yes', 'y']:
                print("\nThank you for using the Enhanced Stock Valuation App!")
                break


if __name__ == "__main__":
    main()